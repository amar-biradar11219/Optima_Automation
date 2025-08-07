import openpyxl
from openpyxl.reader.excel import load_workbook


class Utils:

    @staticmethod
    def read_data_from_excel(filename, sheet):
        data_list = []
        workbook = load_workbook(filename=filename)
        sheet = workbook[sheet]
        row_count = sheet.max_row
        column_count = sheet.max_column

        for i in range(2, row_count+1):
            row = []
            for j in range(1,column_count+1):
                row.append(sheet.cell(row=i,column=j).value)
            data_list.append(row)
        return data_list
        # workbook = openpyxl.load_workbook(filename)
        # sheet = workbook.active
        # data = []
        # for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip headers
        #     data.append((row[0], row[1]))
        # return data
    # @staticmethod
    # def read_test_data(file_path, sheet_name):
    #     workbook = load_workbook(file_path)
    #     sheet = workbook[sheet_name]
    #     data = []
    #
    #     headers = [cell.value for cell in sheet[1]]
    #
    #     for row in sheet.iter_rows(min_row=2, values_only=True):
    #         data.append(dict(zip(headers, row)))
    #
    #     return data
    #
    # def get_row_count(file, sheetName):
    #     workbook = openpyxl.load_workbook(file)
    #     sheet = workbook[sheetName]
    #     return (sheet.max_row)
    #
    # def get_column_count(file, sheetName):
    #     workbook = openpyxl.load_workbook(file)
    #     sheet = workbook[sheetName]
    #     return (sheet.max_column)
    #
    # def read_data(file, sheetName, rownum, columnno):
    #     workbook = openpyxl.load_workbook(file)
    #     sheet = workbook[sheetName]
    #     return sheet.cell(row=rownum, column=columnno).value
    #
    # def write_data(file, sheetName, rownum, columnno, data):
    #     workbook = openpyxl.load_workbook(file)
    #     sheet = workbook[sheetName]
    #     sheet.cell(row=rownum, column=columnno).value = data
    #     workbook.save(file)






